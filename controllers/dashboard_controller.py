# controllers/dashboard_controller.py
from odoo import http
from odoo.http import request
import json
import logging
import io
import base64
import tempfile
import os
from datetime import datetime
from odoo.exceptions import ValidationError, AccessError
from odoo.tools import DEFAULT_SERVER_DATE_FORMAT

_logger = logging.getLogger(__name__)

class DashboardController(http.Controller):
    
    @http.route('/dashboard_projet/data', type='json', auth='user', methods=['POST'], csrf=False)
    def get_dashboard_data(self, date_debut=None, date_fin=None):
        """Endpoint principal pour récupérer les données du dashboard"""
        try:
            _logger.info(f"Requête dashboard - début: {date_debut}, fin: {date_fin}")
            
            if not request.env.user:
                return self._error_response('Utilisateur non authentifié')
            
            if 'dashboard.projet' not in request.env:
                _logger.error("Modèle 'dashboard.projet' non trouvé")
                return self._default_dashboard_data()
            
            date_debut = self._validate_date(date_debut)
            date_fin = self._validate_date(date_fin)
            
            dashboard_model = request.env['dashboard.projet']
            result = dashboard_model.get_dashboard_data(date_debut, date_fin)
            
            result = self._ensure_valid_response(result)
            
            _logger.info(f"Données dashboard retournées avec succès")
            return result
            
        except AccessError as e:
            _logger.error(f"Erreur d'accès: {str(e)}")
            return self._error_response('Accès refusé', include_default=True)
            
        except ValidationError as e:
            _logger.error(f"Erreur de validation: {str(e)}")
            return self._error_response(f'Erreur de validation: {str(e)}', include_default=True)
            
        except Exception as e:
            _logger.error(f"Erreur inattendue: {str(e)}")
            return self._error_response(f'Erreur serveur: {str(e)}', include_default=True)
    
    @http.route('/dashboard_projet/projet_marge/<int:projet_id>', type='json', auth='user', methods=['POST'], csrf=False)
    def get_projet_marge(self, projet_id, date_debut=None, date_fin=None):
        """Endpoint pour récupérer la marge d'un projet spécifique"""
        try:
            _logger.info(f"Marge projet demandée - projet: {projet_id}, début: {date_debut}, fin: {date_fin}")
            
            if not request.env.user:
                return self._error_response('Utilisateur non authentifié', default_marge=True)
            
            if not projet_id or projet_id <= 0:
                return self._default_marge_data()
            
            if 'dashboard.projet' not in request.env:
                _logger.error("Modèle 'dashboard.projet' non trouvé")
                return self._default_marge_data()
            
            date_debut = self._validate_date(date_debut)
            date_fin = self._validate_date(date_fin)
            
            dashboard_model = request.env['dashboard.projet']
            result = dashboard_model.get_marge_salariale_projet(projet_id, date_debut, date_fin)
            
            result = self._ensure_valid_marge(result)
            
            _logger.info(f"Marge projet retournée avec succès pour le projet {projet_id}")
            return result
            
        except AccessError as e:
            _logger.error(f"Erreur d'accès marge projet: {str(e)}")
            return self._error_response('Accès refusé', default_marge=True)
            
        except Exception as e:
            _logger.error(f"Erreur marge projet {projet_id}: {str(e)}")
            return self._error_response(f'Erreur serveur: {str(e)}', default_marge=True)
    
    @http.route('/dashboard_projet/budget_data', type='json', auth='user', methods=['POST'], csrf=False)
    def get_budget_data(self, date_debut=None, date_fin=None):
        """Endpoint pour récupérer les données budgétaires"""
        try:
            _logger.info(f"Données budget demandées - début: {date_debut}, fin: {date_fin}")
            
            if not request.env.user:
                return self._error_response('Utilisateur non authentifié')
            
            if 'dashboard.projet' not in request.env:
                return self._default_budget_data()
            
            date_debut = self._validate_date(date_debut)
            date_fin = self._validate_date(date_fin)
            
            dashboard_model = request.env['dashboard.projet']
            result = dashboard_model.get_budget_comparison(date_debut, date_fin)
            
            return result
            
        except Exception as e:
            _logger.error(f"Erreur données budget: {str(e)}")
            return self._error_response(f'Erreur serveur: {str(e)}')
    
    @http.route('/dashboard_projet/export', type='http', auth='user', methods=['GET'])
    def export_dashboard(self, date_debut=None, date_fin=None, format='json', **kwargs):
        """Endpoint pour exporter les données du dashboard (JSON/CSV)"""
        try:
            _logger.info(f"Export dashboard demandé - format: {format}")
            
            if not request.env.user:
                return request.make_response("Utilisateur non authentifié", status=401)
            
            if 'dashboard.projet' not in request.env:
                return request.make_response("Modèle dashboard non disponible", status=500)
            
            dashboard_model = request.env['dashboard.projet']
            data = dashboard_model.get_dashboard_data(date_debut, date_fin)
            
            if format.lower() == 'json':
                return self._export_json(data, date_debut, date_fin)
            elif format.lower() == 'csv':
                return self._export_csv(data, date_debut, date_fin)
            else:
                return request.make_response(f"Format '{format}' non supporté", status=400)
            
        except Exception as e:
            _logger.error(f"Erreur export dashboard: {str(e)}")
            return request.make_response(f"Erreur export: {str(e)}", status=500)
    
    @http.route('/dashboard_projet/export_excel', type='json', auth='user', methods=['POST'], csrf=False)
    def export_excel(self, date_debut=None, date_fin=None, include_charts=False, include_budget=False, **kwargs):
        """Export Excel avec données détaillées et graphiques"""
        try:
            _logger.info(f"Export Excel demandé - début: {date_debut}, fin: {date_fin}")
            
            if not request.env.user:
                return {'error': 'Utilisateur non authentifié'}
            
            # Import conditionnel d'openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                from openpyxl.chart import BarChart, PieChart, Reference
            except ImportError:
                return {'error': 'Librairie openpyxl non installée. Installez avec: pip install openpyxl'}
            
            if 'dashboard.projet' not in request.env:
                return {'error': 'Modèle dashboard non disponible'}
            
            date_debut = self._validate_date(date_debut)
            date_fin = self._validate_date(date_fin)
            
            # Récupération des données
            dashboard_model = request.env['dashboard.projet']
            data = dashboard_model.get_dashboard_data(date_debut, date_fin)
            
            # Création du fichier Excel
            wb = openpyxl.Workbook()
            
            # Suppression de la feuille par défaut
            wb.remove(wb.active)
            
            # Feuille de résumé
            self._create_summary_sheet(wb, data, date_debut, date_fin)
            
            # Feuille des projets
            self._create_projects_sheet(wb, data, include_budget)
            
            # Feuille des marges si demandée
            if include_budget:
                self._create_budget_sheet(wb, data)
            
            # Feuille des graphiques si demandée
            if include_charts:
                self._create_charts_sheet(wb, data)
            
            # Sauvegarde du fichier
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Création d'un fichier temporaire pour téléchargement
            filename = f"dashboard_{date_debut}_{date_fin}.xlsx"
            file_data = output.getvalue()
            
            # Création de l'attachment pour téléchargement
            attachment = request.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(file_data),
                'store_fname': filename,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'public': True,
            })
            
            download_url = f'/web/content/{attachment.id}?download=true'
            
            return {
                'url': download_url,
                'filename': filename,
                'success': True
            }
            
        except Exception as e:
            _logger.error(f"Erreur export Excel: {str(e)}")
            return {'error': f'Erreur export Excel: {str(e)}'}
    
    @http.route('/dashboard_projet/export_pdf', type='json', auth='user', methods=['POST'], csrf=False)
    def export_pdf(self, date_debut=None, date_fin=None, include_charts=False, **kwargs):
        """Export PDF avec mise en forme professionnelle"""
        try:
            _logger.info(f"Export PDF demandé - début: {date_debut}, fin: {date_fin}")
            
            if not request.env.user:
                return {'error': 'Utilisateur non authentifié'}
            
            # Import conditionnel de reportlab
            try:
                from reportlab.lib.pagesizes import letter, A4
                from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
                from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
                from reportlab.lib import colors
                from reportlab.lib.units import inch, cm
                from reportlab.graphics.shapes import Drawing
                from reportlab.graphics.charts.barcharts import VerticalBarChart
                from reportlab.graphics.charts.piecharts import Pie
            except ImportError:
                return {'error': 'Librairie reportlab non installée. Installez avec: pip install reportlab'}
            
            if 'dashboard.projet' not in request.env:
                return {'error': 'Modèle dashboard non disponible'}
            
            date_debut = self._validate_date(date_debut)
            date_fin = self._validate_date(date_fin)
            
            # Récupération des données
            dashboard_model = request.env['dashboard.projet']
            data = dashboard_model.get_dashboard_data(date_debut, date_fin)
            
            # Création du PDF
            output = io.BytesIO()
            doc = SimpleDocTemplate(
                output,
                pagesize=A4,
                rightMargin=2*cm,
                leftMargin=2*cm,
                topMargin=2*cm,
                bottomMargin=2*cm
            )
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                spaceAfter=30,
                textColor=colors.HexColor('#2c3e50'),
                alignment=1  # Center
            )
            
            subtitle_style = ParagraphStyle(
                'CustomSubtitle',
                parent=styles['Heading2'],
                fontSize=16,
                spaceAfter=20,
                textColor=colors.HexColor('#34495e')
            )
            
            # Construction du contenu PDF
            story = []
            
            # Titre principal
            story.append(Paragraph("Tableau de Bord Projets", title_style))
            story.append(Paragraph(f"Période du {date_debut or 'Non spécifiée'} au {date_fin or 'Non spécifiée'}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # Résumé exécutif
            self._add_executive_summary(story, data, styles, subtitle_style)
            
            # Métriques principales
            self._add_metrics_section(story, data, styles, subtitle_style)
            
            # Tableau des projets
            self._add_projects_table(story, data, styles, subtitle_style)
            
            # Graphiques si demandés
            if include_charts:
                story.append(PageBreak())
                self._add_charts_section(story, data, styles, subtitle_style)
            
            # Footer avec informations
            story.append(Spacer(1, 30))
            story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles['Normal']))
            
            # Génération du PDF
            doc.build(story)
            output.seek(0)
            
            # Création d'un fichier temporaire pour téléchargement
            filename = f"dashboard_{date_debut}_{date_fin}.pdf"
            file_data = output.getvalue()
            
            # Création de l'attachment pour téléchargement
            attachment = request.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(file_data),
                'store_fname': filename,
                'mimetype': 'application/pdf',
                'public': True,
            })
            
            download_url = f'/web/content/{attachment.id}?download=true'
            
            return {
                'url': download_url,
                'filename': filename,
                'success': True
            }
            
        except Exception as e:
            _logger.error(f"Erreur export PDF: {str(e)}")
            return {'error': f'Erreur export PDF: {str(e)}'}
    
    @http.route('/dashboard_projet/test', type='json', auth='user', methods=['POST'], csrf=False)
    def test_dashboard(self):
        """Test endpoint pour vérifier le fonctionnement du dashboard"""
        try:
            models_available = {
                'dashboard.projet': 'dashboard.projet' in request.env,
                'project.project': 'project.project' in request.env,
                'account.move': 'account.move' in request.env,
                'account.move.line': 'account.move.line' in request.env,
                'sale.order': 'sale.order' in request.env,
                'account.analytic.line': 'account.analytic.line' in request.env,
                'hr.employee': 'hr.employee' in request.env,
            }
            
            db_test = {
                'can_read_projects': False,
                'can_read_invoices': False,
                'can_read_timesheets': False
            }
            
            try:
                if models_available['project.project']:
                    project_count = request.env['project.project'].search_count([])
                    db_test['can_read_projects'] = True
                    db_test['project_count'] = project_count
            except:
                pass
            
            try:
                if models_available['account.move']:
                    invoice_count = request.env['account.move'].search_count([('move_type', '=', 'out_invoice')])
                    db_test['can_read_invoices'] = True
                    db_test['invoice_count'] = invoice_count
            except:
                pass
            
            try:
                if models_available['account.analytic.line']:
                    timesheet_count = request.env['account.analytic.line'].search_count([])
                    db_test['can_read_timesheets'] = True
                    db_test['timesheet_count'] = timesheet_count
            except:
                pass
            
            return {
                'status': 'success',
                'message': 'Dashboard controller opérationnel',
                'user': request.env.user.name,
                'user_id': request.env.user.id,
                'company': request.env.company.name,
                'models_available': models_available,
                'database_test': db_test,
                'timestamp': datetime.now().isoformat()
            }
            
        except Exception as e:
            return {
                'status': 'error',
                'message': str(e),
                'timestamp': datetime.now().isoformat()
            }
    
    @http.route('/dashboard_projet/health', type='json', auth='user', methods=['POST'], csrf=False)
    def health_check(self):
        """Endpoint de vérification de santé du dashboard"""
        try:
            health_status = {
                'status': 'healthy',
                'checks': {
                    'database': 'ok',
                    'models': 'ok',
                    'permissions': 'ok',
                    'calculations': 'ok'
                },
                'details': {},
                'timestamp': datetime.now().isoformat()
            }
            
            # Test de connexion base de données
            try:
                request.env.cr.execute("SELECT 1")
                health_status['details']['database'] = 'Connexion DB active'
            except Exception as e:
                health_status['checks']['database'] = 'error'
                health_status['details']['database'] = str(e)
                health_status['status'] = 'degraded'
            
            # Test des modèles requis
            required_models = ['dashboard.projet', 'project.project']
            missing_models = []
            for model in required_models:
                if model not in request.env:
                    missing_models.append(model)
            
            if missing_models:
                health_status['checks']['models'] = 'error'
                health_status['details']['missing_models'] = missing_models
                health_status['status'] = 'unhealthy'
            
            # Test des permissions
            try:
                if 'project.project' in request.env:
                    request.env['project.project'].check_access_rights('read')
                health_status['details']['permissions'] = 'Accès lecture OK'
            except Exception as e:
                health_status['checks']['permissions'] = 'warning'
                health_status['details']['permissions'] = str(e)
                health_status['status'] = 'degraded'
            
            # Test calculs de base
            try:
                if 'dashboard.projet' in request.env:
                    dashboard = request.env['dashboard.projet']
                    test_ca = dashboard.get_chiffre_affaires()
                    health_status['details']['test_calculation'] = f'CA test: {test_ca}'
            except Exception as e:
                health_status['checks']['calculations'] = 'warning'
                health_status['details']['calculation_error'] = str(e)
                if health_status['status'] == 'healthy':
                    health_status['status'] = 'degraded'
            
            return health_status
            
        except Exception as e:
            return {
                'status': 'error',
                'message': f'Health check failed: {str(e)}',
                'timestamp': datetime.now().isoformat()
            }
    
    # ===== MÉTHODES POUR EXCEL =====
    
    def _create_summary_sheet(self, wb, data, date_debut, date_fin):
        """Création de la feuille de résumé"""
        ws = wb.create_sheet(title="Résumé")
        
        # Import des styles
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Styles
        title_font = Font(size=16, bold=True, color="2C3E50")
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        # Titre
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Tableau de Bord Projets - {date_debut or ''} au {date_fin or ''}"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Métriques principales
        metrics = [
            ["Métrique", "Valeur", "Unité", "Commentaire"],
            ["Chiffre d'Affaires", data.get('chiffre_affaires', 0), "€", "CA total période"],
            ["Nombre de Projets", len(data.get('projets', [])), "projets", "Projets actifs"],
            ["CA Total Admin", data.get('marge_administrative', {}).get('ca_total', 0), "€", "CA pour calcul marge admin"],
            ["Coût Admin", data.get('marge_administrative', {}).get('cout_admin', 0), "€", "Coûts administratifs"],
            ["Marge Admin", data.get('marge_administrative', {}).get('marge_admin', 0), "€", "Marge administrative"],
            ["Taux Marge Admin", data.get('marge_administrative', {}).get('taux_marge_admin', 0), "%", "Pourcentage marge admin"]
        ]
        
        for row, (label, value, unit, comment) in enumerate(metrics, start=3):
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'C{row}'] = unit
            ws[f'D{row}'] = comment
            
            if row == 3:  # Header
                for col in 'ABCD':
                    ws[f'{col}{row}'].font = header_font
                    ws[f'{col}{row}'].fill = header_fill
                    ws[f'{col}{row}'].alignment = Alignment(horizontal="center")
        
        # Ajustement des colonnes
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 25
    
    def _create_projects_sheet(self, wb, data, include_budget):
        """Création de la feuille des projets"""
        ws = wb.create_sheet(title="Projets")
        
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        
        # Headers
        headers = ["ID", "Nom", "CA", "Personnel", "Heures", "Statut"]
        if include_budget:
            headers.extend(["Budget Prévu", "Budget Consommé", "Écart %"])
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Données projets
        for row, projet in enumerate(data.get('projets', []), start=2):
            ws.cell(row=row, column=1, value=projet.get('id'))
            ws.cell(row=row, column=2, value=projet.get('name'))
            ws.cell(row=row, column=3, value=projet.get('ca', 0))
            ws.cell(row=row, column=4, value=projet.get('nb_personnes', 0))
            ws.cell(row=row, column=5, value=projet.get('heures', 0))
            ws.cell(row=row, column=6, value=projet.get('stage', ''))
            
            if include_budget:
                budget_prevu = projet.get('budget_prevu', 0)
                budget_consomme = projet.get('budget_consomme', 0)
                ecart = ((budget_consomme / budget_prevu - 1) * 100) if budget_prevu > 0 else 0
                
                ws.cell(row=row, column=7, value=budget_prevu)
                ws.cell(row=row, column=8, value=budget_consomme)
                ws.cell(row=row, column=9, value=ecart)
        
        # Ajustement des colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
    
    def _create_budget_sheet(self, wb, data):
        """Création de la feuille budget"""
        ws = wb.create_sheet(title="Analyse Budget")
        
        # Implémentation similaire aux autres feuilles
        # Création de tableaux d'analyse budgétaire
        pass
    
    def _create_charts_sheet(self, wb, data):
        """Création de la feuille graphiques"""
        ws = wb.create_sheet(title="Graphiques")
        
        # Ajout de graphiques Excel natifs
        # Implémentation des graphiques avec openpyxl
        pass
    
    # ===== MÉTHODES POUR PDF =====
    
    def _add_executive_summary(self, story, data, styles, subtitle_style):
        """Ajout du résumé exécutif"""
        from reportlab.platypus import Paragraph, Spacer
        
        story.append(Paragraph("Résumé Exécutif", subtitle_style))
        
        ca = data.get('chiffre_affaires', 0)
        nb_projets = len(data.get('projets', []))
        marge_admin = data.get('marge_administrative', {})
        
        summary_text = f"""
        <b>Chiffre d'affaires:</b> {ca:,.0f} €<br/>
        <b>Nombre de projets actifs:</b> {nb_projets}<br/>
        <b>Marge administrative:</b> {marge_admin.get('taux_marge_admin', 0):.1f}%<br/>
        <b>Marge nette:</b> {marge_admin.get('marge_admin', 0):,.0f} €
        """
        
        story.append(Paragraph(summary_text, styles['Normal']))
        story.append(Spacer(1, 20))
    
    def _add_metrics_section(self, story, data, styles, subtitle_style):
        """Ajout de la section métriques"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
        from reportlab.lib import colors
        
        story.append(Paragraph("Métriques Principales", subtitle_style))
        
        marge_admin = data.get('marge_administrative', {})
        
        metrics_data = [
            ['Métrique', 'Valeur'],
            ['CA Total', f"{data.get('chiffre_affaires', 0):,.0f} €"],
            ['Projets Actifs', str(len(data.get('projets', [])))],
            ['CA Admin', f"{marge_admin.get('ca_total', 0):,.0f} €"],
            ['Coût Admin', f"{marge_admin.get('cout_admin', 0):,.0f} €"],
            ['Marge Admin', f"{marge_admin.get('marge_admin', 0):,.0f} €"],
            ['Taux Marge', f"{marge_admin.get('taux_marge_admin', 0):.1f}%"]
        ]
        
        table = Table(metrics_data, colWidths=[4*cm, 4*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    def _add_projects_table(self, story, data, styles, subtitle_style):
        """Ajout du tableau des projets"""
        from reportlab.platypus import Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib import colors
        
        story.append(Paragraph("Détail des Projets", subtitle_style))
        
        projets = data.get('projets', [])
        if not projets:
            story.append(Paragraph("Aucun projet trouvé pour cette période.", styles['Normal']))
            return
        
        # Données du tableau
        table_data = [['ID', 'Nom', 'CA (€)', 'Personnel', 'Heures', 'Statut']]
        
        for projet in projets[:20]:  # Limiter à 20 projets pour éviter débordement
            table_data.append([
                str(projet.get('id', '')),
                projet.get('name', '')[:30] + '...' if len(projet.get('name', '')) > 30 else projet.get('name', ''),
                f"{projet.get('ca', 0):,.0f}",
                str(projet.get('nb_personnes', 0)),
                f"{projet.get('heures', 0):.1f}",
                projet.get('stage', '')[:15]
            ])
        
        table = Table(table_data, colWidths=[1*cm, 5*cm, 2*cm, 2*cm, 2*cm, 2*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495E')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
    
    def _add_charts_section(self, story, data, styles, subtitle_style):
        """Ajout de la section graphiques"""
        from reportlab.platypus import Paragraph, Spacer
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.piecharts import Pie
        from reportlab.lib import colors
        
        story.append(Paragraph("Analyses Graphiques", subtitle_style))
        
        # Graphique des projets par statut
        projets = data.get('projets', [])
        stages = {}
        for projet in projets:
            stage = projet.get('stage', 'Non défini')
            stages[stage] = stages.get(stage, 0) + 1
        
        if stages:
            # Graphique en secteurs
            drawing = Drawing(400, 200)
            pie = Pie()
            pie.x = 50
            pie.y = 50
            pie.width = 300
            pie.height = 100
            pie.data = list(stages.values())
            pie.labels = list(stages.keys())
            pie.slices.strokeWidth = 0.5
            drawing.add(pie)
            story.append(drawing)
        
        story.append(Spacer(1, 20))
    
    # ===== MÉTHODES UTILITAIRES =====
    
    def _export_json(self, data, date_debut, date_fin):
        """Export JSON amélioré"""
        response_data = json.dumps(
            {
                'metadata': {
                    'export_date': datetime.now().isoformat(),
                    'period_start': date_debut,
                    'period_end': date_fin,
                    'version': '2.0'
                },
                'data': data
            }, 
            indent=2, 
            default=self._json_serializer, 
            ensure_ascii=False
        )
        
        headers = [
            ('Content-Type', 'application/json; charset=utf-8'),
            ('Content-Disposition', f'attachment; filename="dashboard_{date_debut}_{date_fin}.json"')
        ]
        
        return request.make_response(response_data, headers=headers)
    
    def _export_csv(self, data, date_debut, date_fin):
        """Export CSV amélioré"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Métadonnées
        writer.writerow(['Export Dashboard Projets'])
        writer.writerow(['Date export', datetime.now().strftime('%d/%m/%Y %H:%M')])
        writer.writerow(['Période', f"{date_debut or 'Non spécifiée'} - {date_fin or 'Non spécifiée'}"])
        writer.writerow([])
        
        # Résumé
        writer.writerow(['RÉSUMÉ FINANCIER'])
        writer.writerow(['Métrique', 'Valeur'])
        writer.writerow(['Chiffre d\'affaires', data.get('chiffre_affaires', 0)])
        writer.writerow(['Nombre de projets', len(data.get('projets', []))])
        
        marge_admin = data.get('marge_administrative', {})
        writer.writerow(['CA Total', marge_admin.get('ca_total', 0)])
        writer.writerow(['Coût Admin', marge_admin.get('cout_admin', 0)])
        writer.writerow(['Marge Admin', marge_admin.get('marge_admin', 0)])
        writer.writerow(['Taux Marge Admin (%)', marge_admin.get('taux_marge_admin', 0)])
        writer.writerow([])
        
        # Détail projets
        writer.writerow(['DÉTAIL DES PROJETS'])
        writer.writerow(['ID', 'Nom', 'CA', 'Personnel', 'Heures', 'Statut', 'Budget Prévu', 'Budget Consommé'])
        
        for projet in data.get('projets', []):
            writer.writerow([
                projet.get('id', ''),
                projet.get('name', ''),
                projet.get('ca', 0),
                projet.get('nb_personnes', 0),
                projet.get('heures', 0),
                projet.get('stage', ''),
                projet.get('budget_prevu', 0),
                projet.get('budget_consomme', 0)
            ])
        
        csv_content = output.getvalue().encode('utf-8-sig')  # BOM pour Excel
        output.close()
        
        headers = [
            ('Content-Type', 'text/csv; charset=utf-8'),
            ('Content-Disposition', f'attachment; filename="dashboard_{date_debut}_{date_fin}.csv"')
        ]
        
        return request.make_response(csv_content, headers=headers)
    
    def _validate_date(self, date_str):
        """Valide et convertit une chaîne de date"""
        if not date_str:
            return None
        try:
            datetime.strptime(date_str, '%Y-%m-%d')
            return date_str
        except ValueError:
            _logger.warning(f"Format de date invalide: {date_str}")
            return None
    
    def _error_response(self, message, include_default=False, default_marge=False):
        """Génère une réponse d'erreur standardisée"""
        response = {'error': message}
        
        if default_marge:
            response.update(self._default_marge_data())
        elif include_default:
            response.update(self._default_dashboard_data())
        
        return response
    
    def _default_dashboard_data(self):
        """Données par défaut pour le dashboard"""
        return {
            'chiffre_affaires': 0,
            'projets': [],
            'marge_administrative': {
                'ca_total': 0,
                'cout_admin': 0,
                'marge_admin': 0,
                'taux_marge_admin': 0
            },
            'budget_comparison': {
                'budget_total': 0,
                'budget_consomme': 0,
                'ecart_budget': 0
            }
        }
    
    def _default_marge_data(self):
        """Données par défaut pour la marge"""
        return {
            'revenus': 0,
            'cout_salarial': 0,
            'marge': 0,
            'taux_marge': 0
        }
    
    def _default_budget_data(self):
        """Données par défaut pour le budget"""
        return {
            'budget_total': 0,
            'budget_consomme': 0,
            'ecart_budget': 0,
            'projets_budget': []
        }
    
    def _ensure_valid_response(self, data):
        """S'assure que la réponse a une structure valide"""
        if not isinstance(data, dict):
            return self._default_dashboard_data()
        
        data['chiffre_affaires'] = data.get('chiffre_affaires', 0) or 0
        
        if not isinstance(data.get('projets'), list):
            data['projets'] = []
        
        marge_admin = data.get('marge_administrative', {})
        if not isinstance(marge_admin, dict):
            marge_admin = {}
        
        data['marge_administrative'] = {
            'ca_total': marge_admin.get('ca_total', 0) or 0,
            'cout_admin': marge_admin.get('cout_admin', 0) or 0,
            'marge_admin': marge_admin.get('marge_admin', 0) or 0,
            'taux_marge_admin': marge_admin.get('taux_marge_admin', 0) or 0
        }
        
        # Ajout des données budget si disponibles
        budget_data = data.get('budget_comparison', {})
        if not isinstance(budget_data, dict):
            budget_data = {}
        
        data['budget_comparison'] = {
            'budget_total': budget_data.get('budget_total', 0) or 0,
            'budget_consomme': budget_data.get('budget_consomme', 0) or 0,
            'ecart_budget': budget_data.get('ecart_budget', 0) or 0
        }
        
        return data
    
    def _ensure_valid_marge(self, data):
        """S'assure que les données de marge sont valides"""
        if not isinstance(data, dict):
            return self._default_marge_data()
        
        return {
            'revenus': data.get('revenus', 0) or 0,
            'cout_salarial': data.get('cout_salarial', 0) or 0,
            'marge': data.get('marge', 0) or 0,
            'taux_marge': data.get('taux_marge', 0) or 0
        }
    
    def _json_serializer(self, obj):
        """Sérialiseur JSON pour les objets non sérialisables"""
        if hasattr(obj, 'isoformat'):
            return obj.isoformat()
        elif hasattr(obj, '__str__'):
            return str(obj)
        return None